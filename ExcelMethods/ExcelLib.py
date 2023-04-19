import os
import sys
sys.path.append('..')
from openpyxl import Workbook,load_workbook
from copy import copy
import pandas as pd
from tools import *
from funcInfo import funcInfo
from collections import deque

@funcInfo(
    name="读取区域",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "filename": "文件名",
        "loader":"范围如 A2:D6",
        "sheet":"工作表名",
        "assign":"是否指定最大行，若为True读取输入的区域，若为False读取到选中列最大行"
    },
    note="读取excel指定区域，返回二维列表"
)
def excel_to_list(filename,loader,sheet="Sheet1",assign=False)->list(list()):
    '''task-1
    '''
    try:
        wb = load_workbook(filename=filename)
        ws = wb[sheet]
        front, back = loader.split(':')
        if assign:
            cell_range = ws[front:back]
        else:
            mrow = ws.max_row
            alpha = [c for c in back if c.isalpha()]
            alpha.append(str(mrow))
            reback = ''.join(alpha)
            cell_range = ws[front:reback]
        return [[colume.value for colume in row] for row in cell_range]
    except FileNotFoundError:
        print("文件不存在")
        return
    except KeyError:
        print('工作表不存在')
        return
    except ValueError:
        print("坐标不合法")
        return
    except Exception as e:
        print(e)
        return

@funcInfo(
    name="写入区域",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "lt":"二维列表",
        "filename": "文件名",
        "loader":"某一行的范围A2:D2，定义为一维的输出范围，数据往下追加",
        "sheet":"工作表名",
        "overwrite":"是否覆盖"
    },
    note="二维列表写入到excel指定区域"
)
def list_to_excel(lt,filename,loader,sheet="Sheet1",overwrite=False)->bool:
    ''' task-2
    '''
    try:
        wb = load_workbook(filename=filename)
        ws = wb[sheet]
        front, back = loader.split(':')
        index = 0
        for i, v in enumerate(front):
            if v.isdigit():
                index = i
                break
        startC, startR = front[:index], front[index:]
        for i, v in enumerate(back):
            if v.isdigit():
                index = i
                break
        endC = back[:index]
        ec = scope_to_num(endC) + 1
        sc = scope_to_num(startC)
        sr = int(startR)
        er = sr + len(lt)
    except FileNotFoundError:
        print("文件不存在")
        return False
    except KeyError:
        print('工作表不存在')
        return False
    except ValueError:
        print("坐标不合法")
        return False
    except Exception as e:
        print(e)
        return False
    if len(lt[0])!=ec-sc:    #末列减初列不等于二维的列数
        print("二维列表和表格宽度不匹配")
        return False
    if overwrite==False:
        for i in range(sr,er):
            for j in range(sc,ec):
                if ws.cell(row=i,column=j).value!=None:
                    print("overwrite=True模式下，范围内存在值")
                    return False
    #开始写
    for i in range(sr, er):
        for j in range(sc, ec):
            ir,ic=i-sr,j-sc
            _=ws.cell(row=i,column=j,value=lt[ir][ic])
    try:
        wb.save(filename)
    except:
        print("文件写入失败")
        return False
    return True

@funcInfo(
    name="合并excel",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "dic": "字典 如{'1.excel':'Sheet1'}",
        "output": "文件路径",
        "mode":"备用"
    },
    note="将多个不同文件的Sheet,合并在一个excel的多个Sheet中"
)
def merge_excel(dic,output,mode=None)->bool:
    ''' task-3
    '''
    if os.path.exists(output):
        wb=load_workbook(filename=output)
    else:
        wb=Workbook()
        del wb['Sheet']
    try:
        for file in dic.keys():
            sheet=dic[file]
            twb=load_workbook(filename=file)
            tws=twb[sheet]
            ws=wb.create_sheet()
            mr=tws.max_row
            mc=tws.max_column
            copy_sheet_attributes(tws,ws)
            for i in range(1,mr+1):
                for j in range(1,mc+1):
                    cell=tws.cell(row=i,column=j)
                    target_cell = ws.cell(row=i, column=j)
                    target_cell.value=cell.value
                    target_cell.data_type=cell.data_type
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = copy(cell.number_format)
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)
        wb.save(output)
    except FileNotFoundError:
        print("文件不存在")
        return False
    except KeyError:
        print('工作表不存在')
        return False
    except PermissionError:
        print("权限不足或文件处于打开状态")
        return False
    except Exception as e:
        print(e)
        return False
    return True

@funcInfo(
    name="读取行列",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "filename": "文件路径",
        "loader": "直接输入 列如：C 行如：3 格式字符串型",
        "sheet":"工作表名"
    },
    note="读取指定文件的指定sheet页 ，进行元素列或行（index）提取 "
)
def read_list(filename,loader,sheet="Sheet1"):
    ''' task -4
    '''
    try:
        wb = load_workbook(filename=filename)
        ws=wb[sheet]
    except FileNotFoundError:
        print("文件不存在")
        return
    except KeyError:
        print('工作表不存在')
        return
    if loader.isalpha():
        mr=ws.max_row
        c=scope_to_num(loader)
        return [ws.cell(row=i,column=c).value for i in range(1,mr+1)]
    if loader.isdigit():
        mc=ws.max_column
        r=int(loader)
        return [ws.cell(row=r,column=i).value for i in range(1,mc+1)]

@funcInfo(
    name="数据替换",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "filename": "文件路径",
        "source":"原字符",
        "target":"目标字符",
        "loader": "某一行的范围A2:D2，定义为一维的输出范围，数据往下追加",
        "sheet":"工作表名",
        "assign": "是否指定最大行，若为True读取输入的区域，若为False读取到选中列最大行"
    },
    note="将指定区域的指定字符全部替换"
)
def excel_replace(filename,source,target,loader,sheet='Sheet1',assign=False)->bool:
    ''' task-5
    '''
    try:
        wb = load_workbook(filename=filename)
        ws=wb[sheet]
        pos = loader_to_pos(loader)
    except FileNotFoundError:
        print("文件不存在")
        return False
    except KeyError:
        print('工作表不存在')
        return False
    except Exception as e:
        print('范围输入有误')
        return False
    sr,er=pos[0][0],pos[0][1]
    sc,ec=pos[1][0],pos[1][1]
    if assign==False:
        er=ws.max_row+1
    for i in range(sr,er):
        for j in range(sc,ec):
            cell=ws.cell(row=i,column=j)
            if cell.value==source:
                cell.value=target
    try:
        wb.save(filename)
    except:
        print("权限不足或文件处于打开状态")
        return False
    return True

@funcInfo(
    name="数据筛选",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "filename": "文件路径",
        "column":"输入列",
        "sheet":"工作表名",
        "para":"可变参数限制条件 {"
               "smaller,smallerE,bigger,biggerE"
               "equal,unequal,contain,uncontain"
                "如smallerE=10 表示值小于等于10"

    },
    note="将多个不同文件的Sheet,合并在一个excel的多个Sheet中"
)
def excel_extract(filename,column,sheet='Sheet1' ,**para):
    ''' task-6
    '''
    dic={
        'smaller':smaller,
        'smallerE':smallerE,
        'bigger':bigger,
        'biggerE':biggerE,
        'equal':equal,
        'unequal':unequal,
        'contain':contain,
        'uncontain':uncontain
    }
    wb = load_workbook(filename=filename)
    ws=wb[sheet]
    mr=ws.max_row
    mc=ws.max_column
    ret=[[ws.cell(row=1,column=i).value for i in range(1,mc+1)]]  #默认第一行是标签
    column=scope_to_num(column)
    for i in range(2,mr):
        condition=True
        value=ws.cell(row=i,column=column).value
        for arg in para:
            tmp=para[arg]
            condition &= dic[arg](tmp,value)
        if condition:
            ret.append([ws.cell(row=i,column=j).value for j in range(1,mc+1)])
    return ret

@funcInfo(
    name="数据透视",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "filename": "文件路径",
        "loader":"某一行的范围A2:D2，定义为一维的输出范围，数据往下追加",
        "group":"pandas中groupby用法",
        "agg":"pandas中agg用法",
        "target":"输出路径",
        "sheet":"工作表",
        "assign":"是否指定最大行，若为True读取输入的区域，若为False读取到选中列最大行"
    },
    note="pandas的groupby和aggregate封装"
)
def excel_pivot(filename,loader,group,agg,target,sheet='Sheet1',assign=False)->bool:
    '''  task-7
    '''
    try:
        wb = load_workbook(filename=filename)
        ws = wb[sheet]
        pos=loader_to_pos(loader)
    except FileNotFoundError:
        print("文件不存在")
        return False
    except KeyError:
        print('工作表不存在')
        return False
    except Exception as e:
        print('范围输入有误')
        return False
    sr,er=pos[0][0],pos[0][1]
    sc,ec=pos[1][0],pos[1][1]
    if assign==False:
        er=ws.max_row+1
    data=[]
    column=[ws.cell(row=sr,column=i).value for i in range(sc,ec)]
    for i in range(sr+1,er):
        temp = []
        for j in range(sc,ec):
            temp.append(ws.cell(row=i,column=j).value)
        data.append(temp)
    df = pd.DataFrame(data, columns=column)
    try:
        df1=df.groupby(group).agg(agg)
        writer = pd.ExcelWriter(target, engine="openpyxl")
        df1 = pd.DataFrame(df1, columns=column)
        df1.drop([group], axis=1, inplace=True)
        df1.to_excel(writer, index=True)
        writer.close()
    except KeyError:
        print("分组过程出错")
        return False
    except AttributeError:
        print("聚合过程出错")
        return False
    except PermissionError:
        print("没有写入权限可能是打开文件未关闭")
        return False
    return True

@funcInfo(
    name="数据拆分",
    methodPath="MRobotPackage.ExcelMethods.ExcelLib",
    argsNote={
        "file": "文件路径",
        "batch":"拆分条数",
        "target":"输出路径",
        "hasHead":"是否有表头"
    },
    note="将巨大的excel文件按条数拆成 xx-1,xx-2"
)
def csv_split(file,batch,target,hasHead=False)->bool:
    '''test-8
    '''
    try:
        filename=os.path.basename(file).split('.')[0]
        row ,count=0,1
        buffer=''
        with open(file) as f:
            if hasHead:
                data=f.read(256)      # 假设标题小于256字节
                lt=data.split('\n')
                head=lt[0]
                buffer+='\n'.join(lt[1:])
            while True:
                if row==0:
                    newfile=target+filename+'-'+str(count)+'.csv'
                    out=open(newfile,"w")
                    if hasHead:
                        out.write(head+'\n')
                data=f.read(2048)
                if not data:
                    break
                buffer+=data
                dq = deque(buffer.split('\n'))
                buffer = ''
                while len(dq) != 1:
                    wd = dq.popleft()
                    out.write(wd + '\n')
                    row += 1
                    if row == batch:
                        row = 0
                        count += 1
                        out.close()
                        newfile = target + filename + '-' + str(count) + '.csv'
                        out = open(newfile, "a+")
                        if hasHead:
                            out.write(head + '\n')
                buffer+=dq.popleft()
            if buffer != "":
                out.write(buffer)
                out.close()
    except FileNotFoundError:
        print('文件路径输入有误')
        return False
    except Exception as e:
        print(e)
        return False
    return True
